import os
import time
import argparse
from typing import List, Optional
from openpyxl import load_workbook
import importlib.util


def _load_module(path: str):
    spec = importlib.util.spec_from_file_location(os.path.splitext(os.path.basename(path))[0], path)
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return mod


def _pick(headers: List[str], cands: List[str]) -> Optional[int]:
    for i, h in enumerate(headers, start=1):
        for c in cands:
            if h == c or h.startswith(c) or (c in h and len(h) <= 40):
                return i
    return None


def _ensure_col(ws, headers: List[str], name: str) -> int:
    if name in headers:
        return headers.index(name) + 1
    ws.cell(row=1, column=ws.max_column + 1).value = name
    headers.append(name)
    return ws.max_column


def _format_contexts(ctxs: List[dict], top_k: int = 3) -> str:
    lines = []
    for d in ctxs[:top_k]:
        page = d.get('metadata', {}).get('page')
        text = d.get('text', '')
        lines.append(f"[page:{page}] {text}")
    return "\n---\n".join(lines)


def process_batch(start_row: int, count: int, retries: int, sleep_ms: int, post_sleep_ms: int):
    in_path = r'd:\RagProject\data\test.xlsx'
    out_path = r'd:\RagProject\data\test copy.xlsx'
    qmod = _load_module(os.path.join(os.path.dirname(__file__), '3_query.py'))
    wb = load_workbook(in_path, data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=i).value or '').strip() for i in range(1, ws.max_column + 1)]
    q_col = _pick(headers, ['问题', '问句', 'query', '问题描述', '标题'])
    if q_col is None:
        raise RuntimeError('未找到问题列')
    ans_col = _pick(headers, ['大模型回答']) or _ensure_col(ws, headers, '大模型回答')
    ref_col = _pick(headers, ['参考问答对']) or _ensure_col(ws, headers, '参考问答对')
    flag_col = _pick(headers, ['相似度是否低于0.3']) or _ensure_col(ws, headers, '相似度是否低于0.3（若是的话填是，不是的话可以不填）')

    end_row = ws.max_row
    processed = 0
    r = max(start_row, 2)
    while r <= end_row and processed < count:
        q = str(ws.cell(row=r, column=q_col).value or '').strip()
        if q:
            attempt = 0
            result = None
            while attempt <= retries:
                try:
                    result = qmod.query_rag(q, image_paths=None, top_k=50, rerank_k=10)
                    if isinstance(result.get('answer'), str) and result.get('contexts'):
                        break
                except Exception:
                    pass
                time.sleep((sleep_ms / 1000.0) * (2 ** attempt))
                attempt += 1

            if result:
                ws.cell(row=r, column=ans_col).value = result.get('answer')
                ctxs = result.get('contexts', [])
                ws.cell(row=r, column=ref_col).value = _format_contexts(ctxs)
                taotalscore = float(ctxs[0].get('final_score', 0.0)) if ctxs else 0.0
                ws.cell(row=r, column=flag_col).value = '是' if taotalscore < 0.3 else ''
                time.sleep(post_sleep_ms / 1000.0)
                processed += 1
        r += 1

    def _safe_save():
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        try:
            wb.save(out_path)
            return out_path
        except PermissionError:
            ts = time.strftime('%Y%m%d_%H%M%S')
            alt = os.path.join(os.path.dirname(out_path), f"test copy_{ts}.xlsx")
            wb.save(alt)
            return alt

    saved = _safe_save()
    print(f"已保存到: {saved}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--start', type=int, default=2)
    parser.add_argument('--count', type=int, default=10)
    parser.add_argument('--retries', type=int, default=3)
    parser.add_argument('--sleep_ms', type=int, default=1000)
    parser.add_argument('--post_sleep_ms', type=int, default=300)
    args = parser.parse_args()
    process_batch(args.start, args.count, args.retries, args.sleep_ms, args.post_sleep_ms)


if __name__ == '__main__':
    main()
